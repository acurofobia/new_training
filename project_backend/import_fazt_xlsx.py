import sqlite3
import os
import shutil
import openpyxl
from datetime import datetime

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, 'instance', 'questions.db')
XLSX_DIR = os.path.dirname(BASE_DIR)

timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
NEW_DB_PATH = os.path.join(BASE_DIR, 'instance', f'questions_{timestamp}.db')

FILES = [
    ('Категория_3_250.xlsx', 3),
    ('Категория_4_250.xlsx', 4),
    ('Категория_5_250.xlsx', 5),
    ('Категория_6_250.xlsx', 6),
    ('Категория_7_250.xlsx', 7),
    ('Категория_8_250.xlsx', 8),
]

ANSWER_MARKER = '►'


def get_rgb(cell):
    try:
        color = cell.font.color
        if color and color.type == 'rgb':
            return color.rgb
    except Exception:
        pass
    return None


def is_red_cell(cell):
    return get_rgb(cell) == 'FFFF0000'


def clean(text):
    if text is None:
        return ''
    return str(text).replace('\xa0', ' ').strip()


def parse_xlsx(filepath):
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    questions = []
    current_q = None

    for row in range(2, ws.max_row + 1):  # skip header row
        col1_val = clean(ws.cell(row, 1).value)
        col2_cell = ws.cell(row, 2)
        col2_val = clean(col2_cell.value)

        if not col1_val and not col2_val:
            continue

        if col1_val == ANSWER_MARKER:
            if current_q is not None and col2_val:
                col1_cell = ws.cell(row, 1)
                correct = is_red_cell(col2_cell) or is_red_cell(col1_cell)
                current_q['answers'].append({
                    'text': col2_val,
                    'points': 1 if correct else 0
                })
        elif col1_val.isdigit() and col2_val:
            if current_q is not None:
                questions.append(current_q)
            current_q = {'number': int(col1_val), 'question': col2_val, 'answers': []}

    if current_q is not None:
        questions.append(current_q)

    return questions


def validate(questions, filename):
    issues = []
    for q in questions:
        correct = sum(a['points'] for a in q['answers'])
        if correct != 1:
            issues.append(f"  Q{q['number']}: {correct} правильных ответов")
        if len(q['answers']) == 0:
            issues.append(f"  Q{q['number']}: нет ответов")
    if issues:
        print(f"[WARN] Проблемы в {filename}:")
        for issue in issues:
            print(issue)
    return len(issues) == 0


def import_to_db(questions, org, category):
    conn = sqlite3.connect(NEW_DB_PATH)
    cur = conn.cursor()

    cur.execute(
        "DELETE FROM answers WHERE question_id IN (SELECT id FROM questions WHERE org=? AND category=?)",
        (org, category)
    )
    cur.execute("DELETE FROM questions WHERE org=? AND category=?", (org, category))

    for q in questions:
        cur.execute(
            "INSERT INTO questions (org, category, question, question_number) VALUES (?, ?, ?, ?)",
            (org, category, q['question'], q['number'])
        )
        q_id = cur.lastrowid
        for ans in q['answers']:
            cur.execute(
                "INSERT INTO answers (question_id, org, points, answer) VALUES (?, ?, ?, ?)",
                (q_id, org, ans['points'], ans['text'])
            )

    conn.commit()
    conn.close()


def main():
    print(f"Копирование БД: {DB_PATH} → {NEW_DB_PATH}")
    shutil.copy2(DB_PATH, NEW_DB_PATH)

    for filename, category in FILES:
        filepath = os.path.join(XLSX_DIR, filename)
        print(f"\nОбработка {filename} (fazt, категория {category})...")

        questions = parse_xlsx(filepath)
        print(f"  Найдено вопросов: {len(questions)}")
        print(f"  Найдено ответов: {sum(len(q['answers']) for q in questions)}")
        print(f"  Правильных ответов: {sum(a['points'] for q in questions for a in q['answers'])}")

        ok = validate(questions, filename)
        if not ok:
            print("  [!] Продолжаем несмотря на предупреждения")

        import_to_db(questions, 'fazt', category)
        print(f"  Импорт завершён")

    print("\n=== Проверка БД ===")
    conn = sqlite3.connect(NEW_DB_PATH)
    cur = conn.cursor()
    for category in [3, 4, 5, 6, 7, 8]:
        count_q = cur.execute("SELECT count(*) FROM questions WHERE org='fazt' AND category=?", (category,)).fetchone()[0]
        count_a = cur.execute(
            "SELECT count(*) FROM answers WHERE question_id IN (SELECT id FROM questions WHERE org='fazt' AND category=?)",
            (category,)
        ).fetchone()[0]
        count_correct = cur.execute(
            "SELECT count(*) FROM answers WHERE points=1 AND question_id IN (SELECT id FROM questions WHERE org='fazt' AND category=?)",
            (category,)
        ).fetchone()[0]
        print(f"  Категория {category}: {count_q} вопросов, {count_a} ответов, {count_correct} правильных")
    conn.close()

    print(f"\nГотово. Новая БД: {NEW_DB_PATH}")
    print("Для деплоя скопируйте её на сервер в ./data/questions.db")


if __name__ == '__main__':
    main()
